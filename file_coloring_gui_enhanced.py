import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import os
import sys

class FileColoringAnalyzer:
    def __init__(self):
        self.file1_path = None
        self.file2_path = None
        
    def select_file1(self):
        """选择文件1（上传计划）"""
        self.file1_path = filedialog.askopenfilename(
            title="选择上传分析依据文件",
            filetypes=[("Excel files", "*.xls *.xlsx"), ("All files", "*.*")]
        )
        return self.file1_path
        
    def select_file2(self):
        """选择文件2（实际上传情况）"""
        self.file2_path = filedialog.askopenfilename(
            title="选择需要分析的文件",
            filetypes=[("Excel files", "*.xls *.xlsx"), ("All files", "*.*")]
        )
        return self.file2_path

def convert_xls_to_xlsx(xls_path, xlsx_path):
    """将.xls文件转换为.xlsx文件"""
    df = pd.read_excel(xls_path)
    df.to_excel(xlsx_path, index=False)
    return xlsx_path

def analyze_and_color_file2_complete(file1_path, file2_path, output_path, progress_callback=None):
    """
    完整版本：分析文件1中的上传计划，然后对文件2的所有行进行颜色标记
    """
    # 读取文件
    try:
        df1 = pd.read_excel(file1_path)
        if progress_callback:
            progress_callback(10, "正在处理上传分析依据文件...")
    except Exception as e:
        raise Exception(f"读取上传分析依据文件时出错: {e}")
    
    # 获取文件1中的计划路径集合
    file1_plans = {}  # 路径 -> 文件1中的行索引和文件名称
    planned_rows = df1[df1['上传计划'].notna()]
    
    # 计算待上传条目数量
    pending_upload_count = len(planned_rows)
    
    for idx, row in planned_rows.iterrows():
        path = row.get('路径', '') if pd.notna(row.get('路径', '')) else ''
        file_name = row.get('文件名称', '') if pd.notna(row.get('文件名称', '')) else ''
        if path:
            path = path.strip().rstrip('/')
            # 保存路径对应的文件1信息（行索引和文件名称）
            file1_plans[path] = {
                'index': idx + 2,  # +2 because Excel is 1-based and first row is header
                'file_name': file_name
            }
    
    if progress_callback:
        progress_callback(20, "正在处理需要分析的文件...")
    
    # 为了获取文件2的列信息，先读取一次
    df2_sample = pd.read_excel(file2_path, nrows=1000)
    
    # 如果文件2是.xls格式，需要先转换为.xlsx格式
    temp_xlsx = None
    if file2_path.lower().endswith('.xls'):
        temp_xlsx = file2_path.replace('.xls', '_temp.xlsx')
        convert_xls_to_xlsx(file2_path, temp_xlsx)
        file2_xlsx_path = temp_xlsx
    else:
        file2_xlsx_path = file2_path
    
    # 加载文件2工作簿
    try:
        wb = load_workbook(file2_xlsx_path)
        ws = wb.active
    except Exception as e:
        raise Exception(f"加载需要分析的文件时出错: {e}")
    
    # 定义颜色
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色 - 完全匹配
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # 橙色 - 路径匹配但文件编号为空
    
    # 获取文件夹列和文件编号列索引
    file2_columns = df2_sample.columns.tolist()
    folder_columns = {}
    file_number_col = None
    
    for i in range(1, 7):
        folder_col_name = f'{i}级文件夹'
        if folder_col_name in file2_columns:
            col_index = file2_columns.index(folder_col_name) + 1
            folder_columns[i] = col_index
    
    # 查找文件编号列
    if '文件编号' in file2_columns:
        file_number_col = file2_columns.index('文件编号') + 1
    
    # 添加"数据来源"列
    source_col_index = len(file2_columns) + 1  # 新增列的索引
    ws.cell(row=1, column=source_col_index).value = "数据来源"
    
    # 处理文件2的所有行
    total_rows = ws.max_row
    yellow_count = 0
    orange_count = 0
    
    # 从第2行开始（第1行是表头）
    for row_idx in range(2, total_rows + 1):
        # 构建路径
        path_parts = []
        for i in range(1, 7):
            if i in folder_columns:
                cell_value = ws.cell(row=row_idx, column=folder_columns[i]).value
                if cell_value and str(cell_value).strip() != '':
                    folder_name = str(cell_value).strip()
                    if folder_name not in ['/', '//', '///']:
                        path_parts.append(folder_name)
        
        path = '/'.join(path_parts) if path_parts else ''
        
        # 检查是否在文件1的计划中
        is_planned = False
        source_info = ""
        matched_plan = None
        
        if path and path in file1_plans:
            is_planned = True
            matched_plan = file1_plans[path]
            source_info = f"文件1第{matched_plan['index']}行: {matched_plan['file_name']}"
        elif path:
            # 部分匹配检查
            for planned_path, plan_info in file1_plans.items():
                if planned_path.startswith(path) or path.startswith(planned_path):
                    is_planned = True
                    matched_plan = plan_info
                    source_info = f"文件1第{plan_info['index']}行: {plan_info['file_name']}"
                    break
        
        if is_planned and path:
            # 检查文件编号是否为空
            file_number_empty = True
            if file_number_col:
                file_number_value = ws.cell(row=row_idx, column=file_number_col).value
                if file_number_value and str(file_number_value).strip() != '':
                    file_number_empty = False
            
            # 设置数据来源列
            ws.cell(row=row_idx, column=source_col_index).value = source_info
            
            # 根据文件编号是否为空来标记颜色
            if file_number_empty:
                # 橙色：路径匹配但文件编号为空
                for col_idx in range(1, ws.max_column):
                    ws.cell(row=row_idx, column=col_idx).fill = orange_fill
                orange_count += 1
            else:
                # 黄色：路径匹配且文件编号不为空
                for col_idx in range(1, ws.max_column):
                    ws.cell(row=row_idx, column=col_idx).fill = yellow_fill
                yellow_count += 1
        
        # 更新进度
        if progress_callback and row_idx % max(1, total_rows // 50) == 0:
            progress_percent = 20 + int(70 * row_idx / total_rows)
            progress_callback(progress_percent, f"正在处理第 {row_idx}/{total_rows} 行...")
    
    if progress_callback:
        progress_callback(90, "正在保存结果...")
    
    # 保存结果
    try:
        wb.save(output_path)
        
        # 清理临时文件
        if temp_xlsx and os.path.exists(temp_xlsx):
            os.remove(temp_xlsx)
            
        if progress_callback:
            progress_callback(100, f"分析完成! 黄色: {yellow_count}行, 橙色: {orange_count}行, 待上传: {pending_upload_count}条")
            
        return True, yellow_count, orange_count, pending_upload_count
    except Exception as e:
        raise Exception(f"保存文件时出错: {e}")

def create_gui():
    """创建GUI界面"""
    root = tk.Tk()
    root.title("文件着色分析工具")
    root.geometry("600x550")
    root.resizable(False, False)
    
    analyzer = FileColoringAnalyzer()
    
    # 标题
    title_label = tk.Label(root, text="文件着色分析工具", font=("Arial", 18, "bold"))
    title_label.pack(pady=20)
    
    # 文件选择框架
    file_frame = tk.Frame(root)
    file_frame.pack(pady=10, fill="x", padx=50)
    
    # 文件1选择（上传分析依据文件）
    file1_frame = tk.Frame(file_frame)
    file1_frame.pack(side=tk.TOP, pady=10, fill="x")
    
    tk.Label(file1_frame, text="上传分析依据文件", font=("Arial", 12)).pack()
    file1_button = tk.Button(file1_frame, text="选择文件", 
                            command=lambda: select_file(1, file1_label, analyzer, analyze_button), 
                            bg="#007bff", fg="white", 
                            font=("Arial", 10), 
                            padx=20, pady=5)
    file1_button.pack(pady=10)
    
    file1_label = tk.Label(file1_frame, text="尚未选择文件", 
                          wraplength=400, 
                          fg="gray")
    file1_label.pack()
    
    # 文件2选择（需要分析的文件）
    file2_frame = tk.Frame(file_frame)
    file2_frame.pack(side=tk.TOP, pady=10, fill="x")
    
    tk.Label(file2_frame, text="需要分析的文件", font=("Arial", 12)).pack()
    file2_button = tk.Button(file2_frame, text="选择文件", 
                            command=lambda: select_file(2, file2_label, analyzer, analyze_button), 
                            bg="#007bff", fg="white", 
                            font=("Arial", 10), 
                            padx=20, pady=5)
    file2_button.pack(pady=10)
    
    file2_label = tk.Label(file2_frame, text="尚未选择文件", 
                          wraplength=400, 
                          fg="gray")
    file2_label.pack()
    
    # 统计信息框架
    stats_frame = tk.Frame(root)
    stats_frame.pack(pady=10, fill="x", padx=50)
    
    yellow_label = tk.Label(stats_frame, text="标黄条目: 0条", fg="orange", font=("Arial", 12))
    yellow_label.pack(side=tk.LEFT, padx=10)
    
    orange_label = tk.Label(stats_frame, text="标橙条目: 0条", fg="orange", font=("Arial", 12))
    orange_label.pack(side=tk.LEFT, padx=10)
    
    pending_label = tk.Label(stats_frame, text="待上传条目: 0条", fg="blue", font=("Arial", 12))
    pending_label.pack(side=tk.LEFT, padx=10)
    
    # 进度条
    progress_frame = tk.Frame(root)
    progress_frame.pack(pady=20, fill="x", padx=50)
    
    progress_label = tk.Label(progress_frame, text="准备就绪")
    progress_label.pack()
    
    progress_bar = ttk.Progressbar(progress_frame, length=500, mode='determinate')
    progress_bar.pack(pady=10)
    
    # 分析按钮
    analyze_button = tk.Button(root, text="开始分析", 
                              command=lambda: analyze_thread(analyzer, root, progress_label, progress_bar, yellow_label, orange_label, pending_label), 
                              bg="#28a745", fg="white", 
                              font=("Arial", 14, "bold"), 
                              padx=30, pady=10,
                              state="disabled")
    analyze_button.pack(pady=30)
    
    # 选择文件的函数
    def select_file(file_num, label, analyzer, button):
        if file_num == 1:
            path = analyzer.select_file1()
            if path:
                label.config(text=os.path.basename(path), fg="black")
        else:
            path = analyzer.select_file2()
            if path:
                label.config(text=os.path.basename(path), fg="black")
        
        # 检查是否两个文件都已选择，如果是则启用分析按钮
        check_and_enable_button(analyzer, button)
    
    def check_and_enable_button(analyzer, button):
        """检查两个文件是否已选择，如果是则启用分析按钮"""
        if analyzer.file1_path and analyzer.file2_path:
            button.config(state="normal")
        else:
            button.config(state="disabled")
    
    # 分析线程函数
    def analyze_thread(analyzer, root, progress_label, progress_bar, yellow_label, orange_label, pending_label):
        def run_analysis():
            try:
                # 检查必要文件是否已选择
                if not analyzer.file1_path or not analyzer.file2_path:
                    messagebox.showerror("错误", "请先选择所有必要的文件")
                    return
                
                # 在开始分析时选择输出文件路径
                output_path = filedialog.asksaveasfilename(
                    title="保存分析结果",
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                    initialfile="分析结果.xlsx"
                )
                
                if not output_path:
                    messagebox.showerror("错误", "请选择输出文件保存位置")
                    return
                
                def update_progress(value, message=""):
                    progress_bar['value'] = value
                    if message:
                        progress_label.config(text=message)
                    root.update()
                
                update_progress(0, "开始分析...")
                
                # 执行分析
                result, yellow_count, orange_count, pending_upload_count = analyze_and_color_file2_complete(
                    analyzer.file1_path, 
                    analyzer.file2_path, 
                    output_path,
                    update_progress
                )
                
                if result:
                    # 更新统计信息
                    yellow_label.config(text=f"标黄条目: {yellow_count}条")
                    orange_label.config(text=f"标橙条目: {orange_count}条")
                    pending_label.config(text=f"待上传条目: {pending_upload_count}条")
                    messagebox.showinfo("完成", f"分析完成!\n结果已保存到: {output_path}\n\n标黄: {yellow_count}条\n标橙: {orange_count}条\n待上传: {pending_upload_count}条")
                else:
                    messagebox.showerror("错误", "分析失败")
                    
            except Exception as e:
                progress_label.config(text=f"处理出错: {str(e)}")
                messagebox.showerror("错误", f"处理过程中出错: {str(e)}")
                
        # 在新线程中运行分析，避免界面冻结
        thread = threading.Thread(target=run_analysis)
        thread.daemon = True
        thread.start()
    
    root.mainloop()

if __name__ == "__main__":
    create_gui()