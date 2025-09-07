import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import sys
import os

def convert_xls_to_xlsx(xls_path, xlsx_path):
    """将.xls文件转换为.xlsx文件"""
    print(f"正在将 {os.path.basename(xls_path)} 转换为xlsx格式...")
    df = pd.read_excel(xls_path)
    df.to_excel(xlsx_path, index=False)
    print(f"转换完成: {xlsx_path}")
    return xlsx_path

def analyze_and_color_file2_complete(file1_path, file2_path, output_path):
    """
    完整版本：分析文件1中的上传计划，然后对文件2的所有行进行颜色标记
    """
    print("正在加载文件...")
    
    # 读取文件
    try:
        df1 = pd.read_excel(file1_path)
        print(f"文件1 ({os.path.basename(file1_path)}) 包含 {len(df1)} 行数据")
        
        # 为了获取文件2的总行数信息，先读取一次
        df2_sample = pd.read_excel(file2_path, nrows=1000)
        print(f"文件2 ({os.path.basename(file2_path)}) 正在处理...")
    except Exception as e:
        print(f"读取文件时出错: {e}")
        return False
    
    # 获取文件1中的计划路径集合
    print("正在处理文件1的计划路径...")
    file1_plans = {}  # 路径 -> 是否有上传计划
    planned_rows = df1[df1['上传计划'].notna()]
    
    for _, row in planned_rows.iterrows():
        path = row.get('路径', '') if pd.notna(row.get('路径', '')) else ''
        if path:
            path = path.strip().rstrip('/')
            file1_plans[path] = True  # 标记为有上传计划
    
    print(f"从文件1中提取了 {len(file1_plans)} 个计划路径")
    
    # 如果文件2是.xls格式，需要先转换为.xlsx格式
    temp_xlsx = None
    if file2_path.lower().endswith('.xls'):
        temp_xlsx = file2_path.replace('.xls', '_temp.xlsx')
        convert_xls_to_xlsx(file2_path, temp_xlsx)
        file2_xlsx_path = temp_xlsx
    else:
        file2_xlsx_path = file2_path
    
    # 加载文件2工作簿
    print("正在加载文件2的工作簿...")
    try:
        wb = load_workbook(file2_xlsx_path)
        ws = wb.active
    except Exception as e:
        print(f"加载文件2时出错: {e}")
        return False
    
    # 定义颜色
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色 - 完全匹配
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")      # 红色 - 路径匹配但文件编号为空
    
    # 获取文件夹列和文件编号列索引
    file2_columns = df2_sample.columns.tolist()
    folder_columns = {}
    file_number_col = None
    
    for i in range(1, 7):
        folder_col_name = f'{i}级文件夹'
        if folder_col_name in file2_columns:
            col_index = file2_columns.index(folder_col_name) + 1
            folder_columns[i] = col_index
    
    # 查找文件编号列
    if '文件编号' in file2_columns:
        file_number_col = file2_columns.index('文件编号') + 1
    
    print(f"找到文件夹列: {folder_columns}")
    print(f"文件编号列索引: {file_number_col}")
    
    # 处理文件2的所有行
    print("正在分析和标记文件2的所有行...")
    yellow_count = 0
    red_count = 0
    total_rows = ws.max_row
    
    # 从第2行开始（第1行是表头）
    for row_idx in range(2, total_rows + 1):
        # 构建路径
        path_parts = []
        for i in range(1, 7):
            if i in folder_columns:
                cell_value = ws.cell(row=row_idx, column=folder_columns[i]).value
                if cell_value and str(cell_value).strip() != '':
                    folder_name = str(cell_value).strip()
                    if folder_name not in ['/', '//', '///']:
                        path_parts.append(folder_name)
        
        path = '/'.join(path_parts) if path_parts else ''
        
        # 检查是否在文件1的计划中
        is_planned = False
        if path and path in file1_plans:
            is_planned = True
        elif path:
            # 部分匹配检查
            for planned_path in file1_plans:
                if planned_path.startswith(path) or path.startswith(planned_path):
                    is_planned = True
                    break
        
        if is_planned and path:
            # 检查文件编号是否为空
            file_number_empty = True
            if file_number_col:
                file_number_value = ws.cell(row=row_idx, column=file_number_col).value
                if file_number_value and str(file_number_value).strip() != '':
                    file_number_empty = False
            
            # 根据文件编号是否为空来标记颜色
            if file_number_empty:
                # 红色：路径匹配但文件编号为空
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = red_fill
                red_count += 1
            else:
                # 黄色：路径匹配且文件编号不为空
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = yellow_fill
                yellow_count += 1
        
        # 显示进度
        if row_idx % 5000 == 0:
            print(f"已处理 {row_idx} 行...")
    
    print(f"\n分析完成:")
    print(f"  文件2总行数: {total_rows - 1}")  # -1 because first row is header
    print(f"  黄色标记行数（已完全匹配）: {yellow_count}")
    print(f"  红色标记行数（路径匹配但文件编号为空）: {red_count}")
    
    # 保存结果
    print("正在保存结果...")
    try:
        wb.save(output_path)
        print(f"\n结果已成功保存: {output_path}")
        
        # 清理临时文件
        if temp_xlsx and os.path.exists(temp_xlsx):
            os.remove(temp_xlsx)
            
        return True
    except Exception as e:
        print(f"保存文件时出错: {e}")
        return False

if __name__ == "__main__":
    print("文件2着色分析工具（完整版）")
    print("=" * 30)
    
    if len(sys.argv) == 4:
        file1_path = sys.argv[1]
        file2_path = sys.argv[2]
        output_path = sys.argv[3]
    else:
        file1_path = input("文件1（上传计划）路径: ").strip()
        file2_path = input("文件2（实际上传情况）路径: ").strip()
        output_path = input("输出文件路径: ").strip()
        
        if not output_path.endswith('.xlsx'):
            output_path += '.xlsx'
    
    if not os.path.exists(file1_path):
        print(f"错误: 文件1不存在: {file1_path}")
        sys.exit(1)
        
    if not os.path.exists(file2_path):
        print(f"错误: 文件2不存在: {file2_path}")
        sys.exit(1)
    
    if analyze_and_color_file2_complete(file1_path, file2_path, output_path):
        print("\n处理完成!")
    else:
        print("\n处理失败!")