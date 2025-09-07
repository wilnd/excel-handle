# analyzer.py
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def convert_xls_to_xlsx(xls_path, xlsx_path):
    """将 .xls 转为 .xlsx（pandas 读取再写出）"""
    df = pd.read_excel(xls_path)
    df.to_excel(xlsx_path, index=False)
    return xlsx_path

def analyze_and_color_file2_complete(file1_path, file2_path, output_path, progress_callback=None):
    """
    分析文件1（上传计划），对文件2（实际）逐行标色：
    - 黄色：路径匹配且“文件编号”不为空
    - 橙色：路径匹配但“文件编号”为空
    另：增加“数据来源”列；返回（ok, yellow_count, orange_count, pending_upload_count）
    """
    try:
        df1 = pd.read_excel(file1_path)
        if progress_callback: progress_callback(10, "正在处理上传分析依据文件...")
    except Exception as e:
        raise Exception(f"读取上传分析依据文件时出错: {e}")

    # 文件1：抽取上传计划
    file1_plans = {}   # path -> {'index', 'file_name'}
    planned_rows = df1[df1['上传计划'].notna()]
    pending_upload_count = len(planned_rows)

    for idx, row in planned_rows.iterrows():
        path = row.get('路径', '') if pd.notna(row.get('路径', '')) else ''
        file_name = row.get('文件名称', '') if pd.notna(row.get('文件名称', '')) else ''
        if path:
            path = str(path).strip().rstrip('/')
            file1_plans[path] = {'index': idx + 2, 'file_name': file_name}

    if progress_callback: progress_callback(20, "正在处理需要分析的文件...")

    # 先抽样读列名（文件2）
    df2_sample = pd.read_excel(file2_path, nrows=1000)

    # 若是 .xls，先转成 .xlsx
    temp_xlsx = None
    file2_xlsx_path = file2_path
    if file2_path.lower().endswith('.xls'):
        temp_xlsx = file2_path.replace('.xls', '_temp.xlsx')
        convert_xls_to_xlsx(file2_path, temp_xlsx)
        file2_xlsx_path = temp_xlsx

    try:
        wb = load_workbook(file2_xlsx_path)
        ws = wb.active
    except Exception as e:
        raise Exception(f"加载需要分析的文件时出错: {e}")

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

    # 列索引
    file2_columns = df2_sample.columns.tolist()
    folder_columns = {}
    file_number_col = None

    for i in range(1, 7):
        name = f"{i}级文件夹"
        if name in file2_columns:
            folder_columns[i] = file2_columns.index(name) + 1

    if '文件编号' in file2_columns:
        file_number_col = file2_columns.index('文件编号') + 1

    # 新增“数据来源”列
    source_col_index = len(file2_columns) + 1
    ws.cell(row=1, column=source_col_index).value = "数据来源"

    total_rows = ws.max_row
    yellow_count = 0
    orange_count = 0

    for row_idx in range(2, total_rows + 1):
        # 拼路径
        parts = []
        for i in range(1, 7):
            if i in folder_columns:
                val = ws.cell(row=row_idx, column=folder_columns[i]).value
                if val and str(val).strip() not in ['/', '//', '///', '']:
                    parts.append(str(val).strip())
        path = '/'.join(parts) if parts else ''

        is_planned = False
        source_info = ""
        matched_plan = None

        if path and path in file1_plans:
            is_planned = True
            matched_plan = file1_plans[path]
            source_info = f"文件1第{matched_plan['index']}行: {matched_plan['file_name']}"
        elif path:
            for planned_path, plan_info in file1_plans.items():
                if planned_path.startswith(path) or path.startswith(planned_path):
                    is_planned = True
                    matched_plan = plan_info
                    source_info = f"文件1第{plan_info['index']}行: {plan_info['file_name']}"
                    break

        if is_planned and path:
            file_number_empty = True
            if file_number_col:
                v = ws.cell(row=row_idx, column=file_number_col).value
                if v and str(v).strip() != '':
                    file_number_empty = False

            ws.cell(row=row_idx, column=source_col_index).value = source_info

            # 修正：必须包含最后一列 → range(1, ws.max_column + 1)
            if file_number_empty:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = orange_fill
                orange_count += 1
            else:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = yellow_fill
                yellow_count += 1

        if progress_callback and row_idx % max(1, total_rows // 50) == 0:
            pct = 20 + int(70 * row_idx / total_rows)
            progress_callback(pct, f"正在处理第 {row_idx}/{total_rows} 行...")

    if progress_callback: progress_callback(90, "正在保存结果...")

    try:
        wb.save(output_path)
        if temp_xlsx and os.path.exists(temp_xlsx):
            os.remove(temp_xlsx)

        if progress_callback:
            progress_callback(100, f"分析完成! 黄色: {yellow_count} 行, 橙色: {orange_count} 行, 待上传: {pending_upload_count} 条")

        return True, yellow_count, orange_count, pending_upload_count
    except Exception as e:
        raise Exception(f"保存文件时出错: {e}")
