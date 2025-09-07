from flask import Flask, render_template, request, send_file, jsonify
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import os
import tempfile
import json
import logging
from datetime import datetime

app = Flask(__name__)

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('web_app.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# 存储生成的文件路径，以便下载
generated_files = {}

# 全局变量存储处理状态
processing_status = {
    "status": "idle",
    "message": "",
    "progress": 0,
    "yellow_count": 0,
    "orange_count": 0,
    "pending_upload_count": 0
}

def convert_xls_to_xlsx(xls_path, xlsx_path):
    """将.xls文件转换为.xlsx文件"""
    df = pd.read_excel(xls_path)
    df.to_excel(xlsx_path, index=False)
    return xlsx_path

def analyze_and_color_file2_complete(file1_path, file2_path, output_path, progress_callback=None):
    """
    完整版本：分析文件1中的上传计划，然后对文件2的所有行进行颜色标记
    """
    # 读取文件
    try:
        df1 = pd.read_excel(file1_path)
        if progress_callback:
            progress_callback(10, "正在处理上传分析依据文件...")
    except Exception as e:
        raise Exception(f"读取上传分析依据文件时出错: {e}")
    
    # 获取文件1中的计划路径集合
    file1_plans = {}  # 路径 -> 文件1中的行索引和文件名称
    planned_rows = df1[df1['上传计划'].notna()]
    
    # 计算待上传条目数量
    pending_upload_count = len(planned_rows)
    
    for idx, row in planned_rows.iterrows():
        path = row.get('路径', '') if pd.notna(row.get('路径', '')) else ''
        file_name = row.get('文件名称', '') if pd.notna(row.get('文件名称', '')) else ''
        if path:
            path = path.strip().rstrip('/')
            # 保存路径对应的文件1信息（行索引和文件名称）
            file1_plans[path] = {
                'index': idx + 2,  # +2 because Excel is 1-based and first row is header
                'file_name': file_name
            }
    
    if progress_callback:
        progress_callback(20, "正在处理需要分析的文件...")
    
    # 为了获取文件2的列信息，先读取一次
    df2_sample = pd.read_excel(file2_path, nrows=1000)
    
    # 如果文件2是.xls格式，需要先转换为.xlsx格式
    temp_xlsx = None
    if file2_path.lower().endswith('.xls'):
        temp_xlsx = file2_path.replace('.xls', '_temp.xlsx')
        convert_xls_to_xlsx(file2_path, temp_xlsx)
        file2_xlsx_path = temp_xlsx
    else:
        file2_xlsx_path = file2_path
    
    # 加载文件2工作簿
    try:
        wb = load_workbook(file2_xlsx_path)
        ws = wb.active
    except Exception as e:
        raise Exception(f"加载需要分析的文件时出错: {e}")
    
    # 定义颜色
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色 - 完全匹配
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # 橙色 - 路径匹配但文件编号为空
    
    # 获取文件夹列和文件编号列索引
    file2_columns = df2_sample.columns.tolist()
    folder_columns = {}
    file_number_col = None
    
    for i in range(1, 7):
        folder_col_name = f'{i}级文件夹'
        if folder_col_name in file2_columns:
            col_index = file2_columns.index(folder_col_name) + 1
            folder_columns[i] = col_index
    
    # 查找文件编号列
    if '文件编号' in file2_columns:
        file_number_col = file2_columns.index('文件编号') + 1
    
    # 添加"数据来源"列
    source_col_index = len(file2_columns) + 1  # 新增列的索引
    ws.cell(row=1, column=source_col_index).value = "数据来源"
    
    # 处理文件2的所有行
    total_rows = ws.max_row
    yellow_count = 0
    orange_count = 0
    
    # 从第2行开始（第1行是表头）
    for row_idx in range(2, total_rows + 1):
        # 构建路径
        path_parts = []
        for i in range(1, 7):
            if i in folder_columns:
                cell_value = ws.cell(row=row_idx, column=folder_columns[i]).value
                if cell_value and str(cell_value).strip() != '':
                    folder_name = str(cell_value).strip()
                    if folder_name not in ['/', '//', '///']:
                        path_parts.append(folder_name)
        
        path = '/'.join(path_parts) if path_parts else ''
        
        # 检查是否在文件1的计划中
        is_planned = False
        source_info = ""
        matched_plan = None
        
        if path and path in file1_plans:
            is_planned = True
            matched_plan = file1_plans[path]
            source_info = f"文件1第{matched_plan['index']}行: {matched_plan['file_name']}"
        elif path:
            # 部分匹配检查
            for planned_path, plan_info in file1_plans.items():
                if planned_path.startswith(path) or path.startswith(planned_path):
                    is_planned = True
                    matched_plan = plan_info
                    source_info = f"文件1第{plan_info['index']}行: {plan_info['file_name']}"
                    break
        
        if is_planned and path:
            # 检查文件编号是否为空
            file_number_empty = True
            if file_number_col:
                file_number_value = ws.cell(row=row_idx, column=file_number_col).value
                if file_number_value and str(file_number_value).strip() != '':
                    file_number_empty = False
            
            # 设置数据来源列
            ws.cell(row=row_idx, column=source_col_index).value = source_info
            
            # 根据文件编号是否为空来标记颜色
            if file_number_empty:
                # 橙色：路径匹配但文件编号为空
                for col_idx in range(1, ws.max_column):
                    ws.cell(row=row_idx, column=col_idx).fill = orange_fill
                orange_count += 1
            else:
                # 黄色：路径匹配且文件编号不为空
                for col_idx in range(1, ws.max_column):
                    ws.cell(row=row_idx, column=col_idx).fill = yellow_fill
                yellow_count += 1
        
        # 更新进度
        if progress_callback and row_idx % max(1, total_rows // 50) == 0:
            progress_percent = 20 + int(70 * row_idx / total_rows)
            progress_callback(progress_percent, f"正在处理第 {row_idx}/{total_rows} 行...")
    
    if progress_callback:
        progress_callback(90, "正在保存结果...")
    
    # 保存结果
    try:
        wb.save(output_path)
        
        # 清理临时文件
        if temp_xlsx and os.path.exists(temp_xlsx):
            os.remove(temp_xlsx)
            
        if progress_callback:
            progress_callback(100, f"分析完成! 黄色: {yellow_count}行, 橙色: {orange_count}行, 待上传: {pending_upload_count}条")
            
        return True, yellow_count, orange_count, pending_upload_count
    except Exception as e:
        raise Exception(f"保存文件时出错: {e}")

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_files():
    global processing_status
    
    try:
        # 获取上传的文件
        file1 = request.files['file1']
        file2 = request.files['file2']
        
        if not file1 or not file2:
            return jsonify({"error": "请上传两个文件"}), 400
        
        # 保存文件到临时位置
        temp_dir = tempfile.mkdtemp()
        file1_path = os.path.join(temp_dir, file1.filename)
        file2_path = os.path.join(temp_dir, file2.filename)
        
        file1.save(file1_path)
        file2.save(file2_path)
        
        # 设置输出文件路径
        output_filename = "分析结果.xlsx"
        output_path = os.path.join(temp_dir, output_filename)
        
        # 定义进度回调函数
        def update_progress(value, message=""):
            global processing_status
            processing_status["progress"] = value
            processing_status["message"] = message
        
        # 执行分析
        processing_status["status"] = "processing"
        result, yellow_count, orange_count, pending_upload_count = analyze_and_color_file2_complete(
            file1_path, 
            file2_path, 
            output_path,
            update_progress
        )
        
        if result:
            processing_status["status"] = "completed"
            processing_status["yellow_count"] = yellow_count
            processing_status["orange_count"] = orange_count
            processing_status["pending_upload_count"] = pending_upload_count
            
            # 生成一个唯一的文件ID
            file_id = datetime.now().strftime("%Y%m%d_%H%M%S")
            generated_files[file_id] = output_path
            
            logger.info(f"分析完成 - 标黄: {yellow_count}, 标橙: {orange_count}, 待上传: {pending_upload_count}")
            
            return jsonify({
                "success": True,
                "yellow_count": yellow_count,
                "orange_count": orange_count,
                "pending_upload_count": pending_upload_count,
                "file_id": file_id
            })
        else:
            processing_status["status"] = "error"
            return jsonify({"error": "分析失败"}), 500
            
    except Exception as e:
        processing_status["status"] = "error"
        return jsonify({"error": f"处理过程中出错: {str(e)}"}), 500

@app.route('/status')
def get_status():
    return jsonify(processing_status)

@app.route('/download/<file_id>')
def download_file(file_id):
    if file_id not in generated_files:
        logger.error(f"尝试下载不存在的文件: {file_id}")
        return jsonify({"error": "文件不存在"}), 404
    
    file_path = generated_files[file_id]
    if not os.path.exists(file_path):
        logger.error(f"文件不存在于磁盘: {file_path}")
        return jsonify({"error": "文件不存在于磁盘"}), 404
    
    logger.info(f"下载文件: {file_path}")
    return send_file(file_path, as_attachment=True, download_name="分析结果.xlsx")

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=9080, debug=True)